"""
Transportation Analytics System
Step 4: Generate Excel Analytics Report (multi-sheet)
"""
import pandas as pd
import numpy as np
from openpyxl import Workbook, load_workbook
from openpyxl.styles import (Font, PatternFill, Alignment, Border, Side,
                               GradientFill, numbers)
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule, DataBar, FormatObject
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.chart.series import SeriesLabel
import warnings; warnings.filterwarnings("ignore")

master = pd.read_csv("data/master_analytics_table.csv", parse_dates=["trip_date"])

# ── Style helpers ─────────────────────────────────────────────────────────────
def header_style(cell, color="1565C0"):
    cell.font      = Font(bold=True, color="FFFFFF", size=11, name="Arial")
    cell.fill      = PatternFill("solid", fgColor=color)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border    = Border(bottom=Side(style="medium", color="FFFFFF"))

def sub_header(cell, color="E3F2FD"):
    cell.font      = Font(bold=True, size=10, name="Arial", color="0D47A1")
    cell.fill      = PatternFill("solid", fgColor=color)
    cell.alignment = Alignment(horizontal="center", vertical="center")

def data_cell(cell, align="center"):
    cell.font      = Font(size=10, name="Arial")
    cell.alignment = Alignment(horizontal=align, vertical="center")
    cell.border    = Border(bottom=Side(style="thin", color="E0E0E0"),
                             right=Side(style="thin", color="E0E0E0"))

def title_cell(ws, row, col, text, color="0D47A1", size=14, span=1):
    c = ws.cell(row=row, column=col, value=text)
    c.font      = Font(bold=True, size=size, color=color, name="Arial")
    c.alignment = Alignment(horizontal="center", vertical="center")
    if span > 1:
        ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col+span-1)

def set_col_widths(ws, widths):
    for col_letter, width in widths.items():
        ws.column_dimensions[col_letter].width = width

wb = Workbook()
wb.remove(wb.active)  # remove default sheet

# ═══════════════════════════════════════════════════════════════════════════════
# SHEET 1: Executive Summary Dashboard
# ═══════════════════════════════════════════════════════════════════════════════
ws1 = wb.create_sheet("Executive Summary")
ws1.sheet_view.showGridLines = False
ws1.row_dimensions[1].height = 45

# Title banner
ws1.merge_cells("A1:L1")
c = ws1["A1"]
c.value      = "🚛  TRANSPORTATION ANALYTICS — EXECUTIVE DASHBOARD"
c.font       = Font(bold=True, size=18, color="FFFFFF", name="Arial")
c.fill       = PatternFill("solid", fgColor="1A237E")
c.alignment  = Alignment(horizontal="center", vertical="center")

# KPI boxes ─ row 3-6
kpis = [
    ("Total Trips",    f"{len(master):,}",   "2196F3", "A"),
    ("Total Distance", f"{master['distance_km'].sum():,.0f} km", "00897B", "C"),
    ("Avg Fuel Eff.",  f"{master['fuel_efficiency_kml'].mean():.2f} km/L", "F9A825", "E"),
    ("On-Time Rate",   f"{(master['delivery_status']=='On Time').mean()*100:.1f}%", "43A047", "G"),
    ("Avg Delay",      f"{master['delay_minutes'].mean():.0f} min", "E53935", "I"),
    ("Avg Cost/km",    f"₹{master['cost_per_km'].mean():.0f}", "6A1B9A", "K"),
]
ws1.row_dimensions[3].height = 15
ws1.row_dimensions[4].height = 30
ws1.row_dimensions[5].height = 22
ws1.row_dimensions[6].height = 12
for label, value, color, col in kpis:
    ws1.merge_cells(f"{col}3:{col}3"); ws1.merge_cells(f"{col}4:{chr(ord(col)+1)}4")
    ws1.merge_cells(f"{col}5:{chr(ord(col)+1)}5"); ws1.merge_cells(f"{col}6:{chr(ord(col)+1)}6")
    top = ws1[f"{col}3"]
    top.fill = PatternFill("solid", fgColor=color); top.value = ""
    val_c = ws1[f"{col}4"]; val_c.value = value
    val_c.font = Font(bold=True, size=16, color=color, name="Arial")
    val_c.alignment = Alignment(horizontal="center", vertical="center")
    lbl_c = ws1[f"{col}5"]; lbl_c.value = label
    lbl_c.font = Font(size=10, color="555555", name="Arial")
    lbl_c.alignment = Alignment(horizontal="center")
    bot = ws1[f"{col}6"]
    bot.fill = PatternFill("solid", fgColor=color); bot.value = ""

# Section: Monthly Summary Table
ws1.row_dimensions[8].height = 20
title_cell(ws1, 8, 1, "Monthly Performance Summary", span=8, size=12)
ws1["A8"].fill = PatternFill("solid", fgColor="E8EAF6")

headers = ["Month","Trips","Distance (km)","Fuel (L)","Avg Eff (km/L)","Delays (min)","Cost (₹)","On-Time %"]
for i, h in enumerate(headers, 1):
    c = ws1.cell(row=9, column=i, value=h)
    header_style(c)

monthly = master.groupby("trip_month").agg(
    trips=("trip_id","count"), dist=("distance_km","sum"),
    fuel=("fuel_consumed_l","sum"), eff=("fuel_efficiency_kml","mean"),
    delay=("delay_minutes","mean"), cost=("total_trip_cost_inr","sum"),
    ontime=("delivery_status", lambda x: (x=="On Time").mean()*100)
).reset_index()
month_names = {1:"January",2:"February",3:"March",4:"April",5:"May",6:"June",
               7:"July",8:"August",9:"September",10:"October",11:"November",12:"December"}
alt_fill = PatternFill("solid", fgColor="F5F5F5")
for r_i, row in monthly.iterrows():
    excel_row = 10 + r_i
    ws1.row_dimensions[excel_row].height = 18
    values = [month_names.get(row["trip_month"],"?"), int(row["trips"]),
              f"{row['dist']:,.0f}", f"{row['fuel']:,.0f}", f"{row['eff']:.2f}",
              f"{row['delay']:.0f}", f"₹{row['cost']:,.0f}", f"{row['ontime']:.1f}%"]
    for c_i, v in enumerate(values, 1):
        cell = ws1.cell(row=excel_row, column=c_i, value=v)
        data_cell(cell)
        if r_i % 2 == 0: cell.fill = alt_fill

set_col_widths(ws1, {"A":14,"B":8,"C":16,"D":12,"E":16,"F":14,"G":16,"H":12,
                       "I":14,"J":16,"K":14,"L":14})

# ═══════════════════════════════════════════════════════════════════════════════
# SHEET 2: Master Analytics Table
# ═══════════════════════════════════════════════════════════════════════════════
ws2 = wb.create_sheet("Master Data")
ws2.sheet_view.showGridLines = False
ws2.freeze_panes = "A3"

ws2.merge_cells("A1:AH1")
c = ws2["A1"]
c.value     = "UNIFIED MASTER ANALYTICS TABLE — All Trips"
c.font      = Font(bold=True, size=14, color="FFFFFF", name="Arial")
c.fill      = PatternFill("solid", fgColor="37474F")
c.alignment = Alignment(horizontal="center", vertical="center")
ws2.row_dimensions[1].height = 30

display_cols = ["trip_id","vehicle_id","driver_name","trip_date","route_name","route_category",
                "distance_km","fuel_consumed_l","fuel_efficiency_kml","fuel_cost_inr",
                "traffic_level","weather","delay_minutes","delivery_status",
                "total_trip_cost_inr","cost_per_km","driver_perf_score",
                "vehicle_type","fuel_type","experience_years"]
sub = master[display_cols].copy()
sub["trip_date"] = sub["trip_date"].dt.strftime("%Y-%m-%d")
sub["fuel_efficiency_kml"] = sub["fuel_efficiency_kml"].round(3)

header_colors = {"trip_id":"37474F","vehicle_id":"1565C0","driver_name":"1B5E20",
                  "route_name":"4A148C","route_category":"4A148C","distance_km":"E65100",
                  "fuel_consumed_l":"BF360C","fuel_efficiency_kml":"BF360C",
                  "fuel_cost_inr":"BF360C","traffic_level":"006064","weather":"006064",
                  "delay_minutes":"B71C1C","delivery_status":"B71C1C",
                  "total_trip_cost_inr":"1A237E","cost_per_km":"1A237E",
                  "driver_perf_score":"33691E","vehicle_type":"0D47A1","fuel_type":"0D47A1",
                  "experience_years":"33691E","trip_date":"546E7A"}
col_labels = {"trip_id":"Trip ID","vehicle_id":"Vehicle","driver_name":"Driver",
               "trip_date":"Date","route_name":"Route","route_category":"Category",
               "distance_km":"Dist (km)","fuel_consumed_l":"Fuel (L)","fuel_efficiency_kml":"Eff (km/L)",
               "fuel_cost_inr":"Fuel Cost (₹)","traffic_level":"Traffic","weather":"Weather",
               "delay_minutes":"Delay (min)","delivery_status":"Status",
               "total_trip_cost_inr":"Total Cost (₹)","cost_per_km":"Cost/km (₹)",
               "driver_perf_score":"Perf Score","vehicle_type":"Veh Type",
               "fuel_type":"Fuel Type","experience_years":"Exp (yrs)"}
for c_i, col in enumerate(display_cols, 1):
    cell = ws2.cell(row=2, column=c_i, value=col_labels.get(col, col))
    cell.font      = Font(bold=True, color="FFFFFF", size=10, name="Arial")
    cell.fill      = PatternFill("solid", fgColor=header_colors.get(col,"455A64"))
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
ws2.row_dimensions[2].height = 28

alt_fill2 = PatternFill("solid", fgColor="FAFAFA")
status_fill = {"On Time":PatternFill("solid",fgColor="E8F5E9"),
               "Minor Delay":PatternFill("solid",fgColor="FFF8E1"),
               "Major Delay":PatternFill("solid",fgColor="FFEBEE")}
for r_i, (_, row) in enumerate(sub.iterrows()):
    excel_row = 3 + r_i
    for c_i, col in enumerate(display_cols, 1):
        cell = ws2.cell(row=excel_row, column=c_i, value=row[col])
        cell.font      = Font(size=9, name="Arial")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        if col == "delivery_status":
            cell.fill = status_fill.get(str(row[col]), PatternFill())
        elif r_i % 2 == 0:
            cell.fill = alt_fill2

ws2.auto_filter.ref = f"A2:{get_column_letter(len(display_cols))}2"
# Column widths
widths2 = {"A":10,"B":10,"C":14,"D":12,"E":18,"F":11,"G":10,"H":9,"I":10,"J":12,
            "K":10,"L":10,"M":11,"N":14,"O":14,"P":11,"Q":11,"R":10,"S":10,"T":9}
for col_letter, width in widths2.items():
    ws2.column_dimensions[col_letter].width = width

# ═══════════════════════════════════════════════════════════════════════════════
# SHEET 3: Driver Leaderboard
# ═══════════════════════════════════════════════════════════════════════════════
ws3 = wb.create_sheet("Driver Leaderboard")
ws3.sheet_view.showGridLines = False

ws3.merge_cells("A1:J1")
c = ws3["A1"]
c.value = "DRIVER PERFORMANCE LEADERBOARD"; c.font = Font(bold=True,size=16,color="FFFFFF",name="Arial")
c.fill = PatternFill("solid",fgColor="1B5E20"); c.alignment = Alignment(horizontal="center",vertical="center")
ws3.row_dimensions[1].height = 38

driver_lb = master.groupby("driver_name").agg(
    total_trips=("trip_id","count"),
    avg_perf_score=("driver_perf_score","mean"),
    avg_efficiency=("fuel_efficiency_kml","mean"),
    avg_delay=("delay_minutes","mean"),
    avg_cost_km=("cost_per_km","mean"),
    safety_rating=("safety_rating","mean"),
    experience_years=("experience_years","mean"),
    on_time_pct=("delivery_status", lambda x: (x=="On Time").mean()*100),
    total_distance=("distance_km","sum"),
).reset_index().sort_values("avg_perf_score",ascending=False).reset_index(drop=True)
driver_lb.insert(0,"rank", range(1, len(driver_lb)+1))

headers3 = ["Rank","Driver","Trips","Perf Score","Fuel Eff (km/L)","Avg Delay (min)","Cost/km (₹)","Safety Rating","On-Time %","Exp (yrs)"]
for i,h in enumerate(headers3,1):
    c = ws3.cell(row=2,column=i,value=h)
    header_style(c,"1B5E20")
ws3.row_dimensions[2].height = 25

medal_colors = {1:"FFD700",2:"C0C0C0",3:"CD7F32"}
for r_i, row_data in driver_lb.iterrows():
    excel_row = 3+r_i
    ws3.row_dimensions[excel_row].height = 20
    values = [int(row_data["rank"]), row_data["driver_name"], int(row_data["total_trips"]),
              f"{row_data['avg_perf_score']:.1f}", f"{row_data['avg_efficiency']:.2f}",
              f"{row_data['avg_delay']:.0f}", f"₹{row_data['avg_cost_km']:.0f}",
              f"{row_data['safety_rating']:.1f}/5.0", f"{row_data['on_time_pct']:.1f}%",
              int(row_data["experience_years"])]
    bg = medal_colors.get(int(row_data["rank"]), "FFFFFF" if r_i%2 else "F1F8E9")
    for c_i,v in enumerate(values,1):
        cell = ws3.cell(row=excel_row,column=c_i,value=v)
        cell.font = Font(size=10,name="Arial",bold=(int(row_data["rank"])<=3))
        cell.alignment = Alignment(horizontal="center",vertical="center")
        cell.fill = PatternFill("solid",fgColor=bg)

set_col_widths(ws3,{"A":7,"B":16,"C":8,"D":12,"E":15,"F":15,"G":12,"H":14,"I":11,"J":10})

# ═══════════════════════════════════════════════════════════════════════════════
# SHEET 4: Route Analysis
# ═══════════════════════════════════════════════════════════════════════════════
ws4 = wb.create_sheet("Route Analysis")
ws4.sheet_view.showGridLines = False

ws4.merge_cells("A1:K1")
c = ws4["A1"]
c.value = "ROUTE COST & EFFICIENCY ANALYSIS"
c.font = Font(bold=True,size=15,color="FFFFFF",name="Arial")
c.fill = PatternFill("solid",fgColor="4A148C"); c.alignment = Alignment(horizontal="center",vertical="center")
ws4.row_dimensions[1].height = 35

route_agg = master.groupby(["route_name","route_category"]).agg(
    trips=("trip_id","count"), avg_dist=("distance_km","mean"),
    avg_fuel_eff=("fuel_efficiency_kml","mean"), avg_delay=("delay_minutes","mean"),
    avg_cost_km=("cost_per_km","mean"), total_cost=("total_trip_cost_inr","sum"),
    major_delays=("delivery_status", lambda x: (x=="Major Delay").sum()),
    avg_difficulty=("road_difficulty","mean"),
).reset_index().sort_values("avg_cost_km",ascending=False)

headers4 = ["Route","Category","Trips","Avg Dist (km)","Avg Fuel Eff","Avg Delay (min)",
             "Avg Cost/km (₹)","Total Cost (₹)","Major Delays","Difficulty","Risk Level"]
for i,h in enumerate(headers4,1):
    c = ws4.cell(row=2,column=i,value=h)
    header_style(c,"4A148C")
ws4.row_dimensions[2].height = 25

avg_cost = route_agg["avg_cost_km"].mean()
for r_i, row_data in route_agg.reset_index(drop=True).iterrows():
    excel_row = 3+r_i
    ws4.row_dimensions[excel_row].height = 20
    cost_km = row_data["avg_cost_km"]
    delay   = row_data["avg_delay"]
    risk    = "🔴 High" if (cost_km > avg_cost*1.2 or delay > 60) else \
              ("🟡 Medium" if (cost_km > avg_cost*0.9 or delay > 30) else "🟢 Low")
    values  = [row_data["route_name"], row_data["route_category"], int(row_data["trips"]),
               f"{row_data['avg_dist']:.0f}", f"{row_data['avg_fuel_eff']:.2f}",
               f"{row_data['avg_delay']:.0f}", f"₹{cost_km:.0f}",
               f"₹{row_data['total_cost']:,.0f}", int(row_data["major_delays"]),
               f"{row_data['avg_difficulty']:.1f}", risk]
    bg = "FFF3E0" if cost_km > avg_cost else ("E8F5E9" if cost_km < avg_cost*0.8 else "FFFFFF")
    for c_i,v in enumerate(values,1):
        cell = ws4.cell(row=excel_row,column=c_i,value=v)
        cell.font = Font(size=10,name="Arial")
        cell.alignment = Alignment(horizontal="center",vertical="center")
        cell.fill = PatternFill("solid",fgColor=bg if r_i%2==0 else "FAFAFA")

set_col_widths(ws4,{"A":20,"B":12,"C":8,"D":14,"E":14,"F":16,"G":16,"H":18,"I":14,"J":12,"K":12})

# ═══════════════════════════════════════════════════════════════════════════════
# SHEET 5: Vehicle Analytics
# ═══════════════════════════════════════════════════════════════════════════════
ws5 = wb.create_sheet("Vehicle Analytics")
ws5.sheet_view.showGridLines = False

ws5.merge_cells("A1:K1")
c = ws5["A1"]
c.value = "VEHICLE PERFORMANCE & MAINTENANCE ANALYTICS"
c.font = Font(bold=True,size=14,color="FFFFFF",name="Arial")
c.fill = PatternFill("solid",fgColor="0D47A1"); c.alignment = Alignment(horizontal="center",vertical="center")
ws5.row_dimensions[1].height = 35

veh_agg = master.drop_duplicates("vehicle_id").merge(
    master.groupby("vehicle_id").agg(
        trips=("trip_id","count"), avg_eff=("fuel_efficiency_kml","mean"),
        avg_delay=("delay_minutes","mean"), avg_cost_km=("cost_per_km","mean"),
        total_km=("distance_km","sum"), total_fuel=("fuel_consumed_l","sum"),
    ).reset_index(), on="vehicle_id"
)[["vehicle_id","vehicle_type","fuel_type","year_mfg","base_km_per_l",
   "total_maint_cost_inr","trips","avg_eff","avg_delay","avg_cost_km","total_km","total_fuel"]].sort_values("avg_eff",ascending=False)

headers5 = ["Vehicle ID","Type","Fuel","Year","Base Eff","Maint Cost (₹)","Trips",
             "Avg Eff (km/L)","Avg Delay","Cost/km (₹)","Total km","Total Fuel (L)"]
for i,h in enumerate(headers5,1):
    c = ws5.cell(row=2,column=i,value=h)
    header_style(c,"0D47A1")
ws5.row_dimensions[2].height = 25

for r_i, row_data in veh_agg.reset_index(drop=True).iterrows():
    excel_row = 3+r_i
    ws5.row_dimensions[excel_row].height = 18
    values = [row_data["vehicle_id"], row_data["vehicle_type"], row_data["fuel_type"],
              int(row_data["year_mfg"]), f"{row_data['base_km_per_l']:.1f}",
              f"₹{row_data['total_maint_cost_inr']:,.0f}", int(row_data["trips"]),
              f"{row_data['avg_eff']:.2f}", f"{row_data['avg_delay']:.0f}",
              f"₹{row_data['avg_cost_km']:.0f}", f"{row_data['total_km']:,.0f}",
              f"{row_data['total_fuel']:,.0f}"]
    for c_i,v in enumerate(values,1):
        cell = ws5.cell(row=excel_row,column=c_i,value=v)
        cell.font = Font(size=10,name="Arial")
        cell.alignment = Alignment(horizontal="center",vertical="center")
        if r_i%2==0: cell.fill = PatternFill("solid",fgColor="E3F2FD")

set_col_widths(ws5,{"A":12,"B":12,"C":10,"D":8,"E":10,"F":16,"G":8,
                     "H":13,"I":11,"J":12,"K":12,"L":14})

# ─── Save ─────────────────────────────────────────────────────────────────────
wb.save("outputs/Transportation_Analytics_Report.xlsx")
print("✅ Excel report saved: outputs/Transportation_Analytics_Report.xlsx")
print(f"   Sheets: {[s.title for s in wb.worksheets]}")
